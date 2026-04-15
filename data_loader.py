"""
환불표.xlsx → Python 데이터 구조 변환 모듈

엑셀 파일을 수정하면 이 모듈이 자동으로 새 데이터를 읽어
refund_calculator.py에 반영됩니다.

Sheet1:
  - B~D열: 기간권 raw data (상품명, 사용일수, 환불금액)
  - N~Q열: 시간권 raw data (상품명, 이용시간, 환불금액)
  - R~S열: 시간권 가격 요약 (상품명, 상품금액)

Sheet2:
  - 확장 기간권 주차별 환불비율표 (2~20주권)
"""
import re
from pathlib import Path
import openpyxl

# 엑셀 파일 경로 (data_loader.py 와 같은 디렉토리)
_XLSX_PATH = Path(__file__).parent / "환불표.xlsx"

# 시간당 차감금액 및 유효기간 (엑셀에 별도 셀 없으므로 여기서 관리)
HOURLY_DEDUCTION: int = 1_500
TIME_PASS_EXPIRY_DAYS: int = 28


def _load_workbook():
    """data_only=True 로 엑셀 로드 (수식 대신 캐시된 값 반환)."""
    if not _XLSX_PATH.exists():
        raise FileNotFoundError(
            f"환불표.xlsx 파일을 찾을 수 없습니다: {_XLSX_PATH}\n"
            "프로젝트 루트에 '환불표.xlsx' 파일을 위치시켜 주세요."
        )
    return openpyxl.load_workbook(_XLSX_PATH, data_only=True)


def load_period_pass_table() -> dict:
    """
    Sheet1 C~E열 raw data에서 기간권 환불 테이블을 구성합니다.
    - C열: 상품명
    - D열: 사용일수
    - E열: 환불금액

    상품금액은 Sheet1 G~H열 요약표에서 읽습니다.

    반환 형식:
    {
        "상품명": {
            "price": int,
            "total_weeks": int,
            "brackets": [(max_day, refund), ...]
        }, ...
    }
    """
    wb = _load_workbook()
    ws = wb["Sheet1"]

    # {상품명: {사용일수: 환불금액}}
    raw: dict[str, dict[int, int]] = {}

    for row in ws.iter_rows(min_row=4, min_col=3, max_col=5, values_only=True):
        name, day, refund = row
        if not isinstance(name, str) or not isinstance(day, int):
            continue
        raw.setdefault(name, {})[day] = int(refund) if isinstance(refund, (int, float)) else 0

    table = {}
    for name, day_map in raw.items():
        week_match = re.search(r"(\d+)주", name)
        if not week_match:
            continue
        total_weeks = int(week_match.group(1))

        # 각 주차 마지막 날의 환불금액을 bracket으로 추출
        brackets = []
        for week in range(1, total_weeks + 1):
            last_day = week * 7
            refund = day_map.get(last_day, 0)
            brackets.append((last_day, int(round(refund))))

        table[name] = {
            "price": 0,  # 아래에서 채움
            "total_weeks": total_weeks,
            "brackets": brackets,
        }

    # Sheet1 G~H열에서 상품금액 보완
    for row in ws.iter_rows(min_row=5, max_row=20, min_col=7, max_col=8, values_only=True):
        name, price = row
        if isinstance(name, str) and name in table and isinstance(price, (int, float)):
            table[name]["price"] = int(round(price))

    wb.close()
    return table


def load_time_pass_prices() -> dict:
    """
    Sheet1 S~T열에서 시간권 가격표를 읽습니다.
    - S열(19): 상품명
    - T열(20): 상품금액

    반환 형식: {"상품명": 가격(int), ...}
    """
    wb = _load_workbook()
    ws = wb["Sheet1"]

    prices = {}
    for row in ws.iter_rows(min_row=2, max_row=20, min_col=19, max_col=20, values_only=True):
        name, price = row
        if isinstance(name, str) and isinstance(price, (int, float)) and price > 0:
            prices[name] = int(round(price))

    wb.close()
    return prices


def load_extended_refund_ratios() -> dict[int, list[float]]:
    """
    Sheet2에서 확장 기간권 주차별 환불비율표를 읽습니다.

    반환 형식: {주수(int): [주차별_비율, ...], ...}
    예: {2: [0.4], 4: [0.7, 0.4, 0.2, 0.0], ...}
    """
    wb = _load_workbook()
    ws = wb["Sheet2"]

    ratios = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not isinstance(name, str):
            continue
        week_match = re.search(r"(\d+)주", name)
        if not week_match:
            continue
        total_weeks = int(week_match.group(1))
        # None 제거 후 0.0 추가 (마지막 분기 환불 불가)
        week_ratios = [v for v in row[1:] if v is not None]
        week_ratios.append(0.0)  # 마지막 주차 이후 환불 불가
        ratios[total_weeks] = week_ratios

    wb.close()
    return ratios


def load_all() -> dict:
    """
    모든 데이터를 한 번에 로드하여 반환합니다.

    반환 형식:
    {
        "period_pass_table": dict,
        "time_pass_prices": dict,
        "extended_ratios": dict,
        "hourly_deduction": int,
        "time_pass_expiry_days": int,
    }
    """
    return {
        "period_pass_table": load_period_pass_table(),
        "time_pass_prices": load_time_pass_prices(),
        "extended_ratios": load_extended_refund_ratios(),
        "hourly_deduction": HOURLY_DEDUCTION,
        "time_pass_expiry_days": TIME_PASS_EXPIRY_DAYS,
    }


if __name__ == "__main__":
    # 로드 확인용 직접 실행
    data = load_all()

    print("=== 기간권 환불 테이블 ===")
    for name, info in data["period_pass_table"].items():
        print(f"  {name}: {info['price']:,}원, {info['total_weeks']}주")
        for max_day, refund in info["brackets"]:
            print(f"    ~{max_day}일: {refund:,}원")

    print("\n=== 시간권 가격표 ===")
    for name, price in data["time_pass_prices"].items():
        print(f"  {name}: {price:,}원")

    print("\n=== 확장 기간권 비율 (2~20주) ===")
    for weeks, ratios in sorted(data["extended_ratios"].items()):
        print(f"  {weeks}주권: {[f'{r*100:.0f}%' for r in ratios]}")
